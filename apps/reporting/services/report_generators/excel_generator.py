"""
Excel report generator implementation.
"""

import logging
from io import BytesIO
from typing import TYPE_CHECKING, List
from uuid import UUID

from openpyxl import Workbook  # type: ignore[reportMissingModuleSource]
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore[reportMissingModuleSource]
from openpyxl.utils import get_column_letter  # type: ignore[reportMissingModuleSource]
from openpyxl.worksheet.datavalidation import DataValidation  # type: ignore[reportMissingModuleSource]

from apps.results_manager.providers import get_results_provider
from apps.review_results.models import SimpleReviewDecision

from . import ReportGenerator

if TYPE_CHECKING:
    from apps.reporting.models import ExportReport

logger = logging.getLogger(__name__)


class ExcelReportGenerator(ReportGenerator):
    """Generate Excel reports for offline backup export."""

    # Agent Grey Brand Colours (from UI Style Guide)
    COLOUR_DEEP_NAVY = "0A2D45"  # Headers, primary text
    COLOUR_TEAL_BLUE = "00A0A0"  # Links, interactive
    COLOUR_AMBER_GOLD = "F6A623"  # Status: Processing/Caution
    COLOUR_OFF_WHITE = "F9F9F7"  # Alternating rows
    COLOUR_COOL_GREY_LIGHT = "E4E6E8"  # Borders, dividers
    COLOUR_COOL_GREY_DARK = "7B8794"  # Secondary text
    COLOUR_STATUS_GREEN = "2E8540"  # Status: Include/Ready
    COLOUR_STATUS_BLUE = "0066FF"  # Primary buttons
    COLOUR_ERROR_RED = "D72638"  # Status: Exclude/Error
    COLOUR_WHITE = "FFFFFF"  # Background

    # Validation choices for dropdowns
    REVIEW_DECISIONS = ["Include", "Exclude", "Maybe"]
    EXCLUSION_REASONS = [
        display for _, display in SimpleReviewDecision.EXCLUSION_REASONS
    ]

    def generate(self, report: "ExportReport", data: dict) -> bytes:
        """Generate Excel report content."""
        try:
            if report.report_type == "offline_backup":
                return self._generate_offline_backup_excel(data)
            else:
                raise ValueError(f"Unsupported report type: {report.report_type}")
        except Exception as e:
            logger.exception(f"Excel generation failed for report {report.id}")
            raise ExcelGenerationError(f"Failed to generate Excel backup: {str(e)}")

    def get_content_type(self) -> str:
        """Get MIME content type for Excel."""
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def get_file_extension(self) -> str:
        """Get file extension for Excel."""
        return "xlsx"

    def _generate_offline_backup_excel(self, data: dict) -> bytes:
        """Generate offline backup Excel workbook with review data and PRISMA statistics."""
        logger.info("Starting offline backup Excel generation")

        wb = Workbook()

        # Get session data
        session_id = data.get("session_id")
        if not session_id:
            raise ValueError("Session ID is required for offline backup generation")

        logger.info(f"Generating offline backup for session {session_id}")

        # Validate session exists and has results
        results_provider = get_results_provider()
        results_count = results_provider.get_results_count(session_id)

        if results_count == 0:
            logger.warning(
                f"No results found for session {session_id} during offline backup generation"
            )
        else:
            logger.info(f"Found {results_count} results for session {session_id}")

        # Create Instructions sheet first
        self._create_instructions_sheet(wb, data)

        # Create Review Results sheet with data validation
        review_sheet = self._create_review_results_sheet(wb, session_id)

        # Create PRISMA Statistics sheet with formulas
        self._create_prisma_statistics_sheet(wb, session_id)

        # Set the Review Results as the active sheet
        wb.active = review_sheet

        logger.info(
            "Successfully created Excel workbook with Instructions, Review Results and PRISMA Statistics sheets"
        )

        return self._save_workbook_to_bytes(wb)

    def _create_review_results_sheet(self, wb: Workbook, session_id: str):
        """Create Review Results sheet with Agent Grey styling and data validation."""
        ws = wb.create_sheet("Review Results")

        # Define headers with optimal column widths
        # Search Date moved to end as metadata-only column
        headers_config = [
            ("Title", 40),
            ("Snippet", 60),
            ("Review Decision", 18),
            ("Exclusion Reason", 35),
            ("Notes", 40),
            ("Search Query(ies)", 35),
            ("URL", 30),
            ("Search Date", 15),
        ]

        # Agent Grey header styling
        header_font = Font(name="Calibri", size=12, bold=True, color=self.COLOUR_WHITE)
        header_fill = PatternFill(
            start_color=self.COLOUR_DEEP_NAVY,
            end_color=self.COLOUR_DEEP_NAVY,
            fill_type="solid",
        )
        header_border = self._create_border()
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Add headers with formatting
        for col, (header, width) in enumerate(headers_config, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = header_alignment
            # Set column width
            ws.column_dimensions[get_column_letter(col)].width = width

        # Get results data
        results_data = self._get_results_data(session_id)

        # Data styling - thin borders
        data_border = self._create_border()
        data_font = Font(name="Calibri", size=11)
        metadata_font = Font(name="Calibri", size=10, color=self.COLOUR_COOL_GREY_DARK)
        link_font = Font(
            name="Calibri", size=11, color=self.COLOUR_TEAL_BLUE, underline="single"
        )

        # Alternating row colours (Agent Grey Off-White and White)
        off_white_fill = PatternFill(
            start_color=self.COLOUR_OFF_WHITE,
            end_color=self.COLOUR_OFF_WHITE,
            fill_type="solid",
        )
        white_fill = PatternFill(
            start_color=self.COLOUR_WHITE,
            end_color=self.COLOUR_WHITE,
            fill_type="solid",
        )

        # Populate data rows with alternating colours and wrapping
        # New column order: Title, Snippet, Review Decision, Exclusion Reason, Notes, Search Query(ies), URL, Search Date
        for row_idx, result_data in enumerate(results_data, start=2):
            # Determine row fill (alternating)
            row_fill = off_white_fill if row_idx % 2 == 0 else white_fill

            # Title (column 1) - clickable hyperlink with embedded URL
            title_cell = ws.cell(row=row_idx, column=1, value=result_data["title"])
            # Embed URL as hyperlink if available
            if result_data["url"]:
                title_cell.hyperlink = result_data["url"]
                # Use teal hyperlink styling for consistency with Agent Grey brand
                title_cell.font = Font(
                    name="Calibri",
                    size=11,
                    color=self.COLOUR_TEAL_BLUE,
                    underline="single",
                )
            else:
                title_cell.font = data_font
            title_cell.fill = row_fill
            title_cell.border = data_border
            title_cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Snippet (column 2) - wrapped text
            snippet_cell = ws.cell(row=row_idx, column=2, value=result_data["snippet"])
            snippet_cell.font = data_font
            snippet_cell.fill = row_fill
            snippet_cell.border = data_border
            snippet_cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Review Decision (column 3) - status colour background
            decision_cell = ws.cell(
                row=row_idx, column=3, value=result_data["decision"]
            )
            decision_cell.font = Font(name="Calibri", size=11, bold=True)
            decision_cell.fill = self._get_decision_fill(
                result_data["decision"], row_fill
            )
            decision_cell.border = data_border
            decision_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Exclusion Reason (column 4) - wrapped text
            reason_cell = ws.cell(
                row=row_idx, column=4, value=result_data["exclusion_reason"]
            )
            reason_cell.font = data_font
            reason_cell.fill = row_fill
            reason_cell.border = data_border
            reason_cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Notes (column 5) - wrapped text
            notes_cell = ws.cell(row=row_idx, column=5, value=result_data["notes"])
            notes_cell.font = data_font
            notes_cell.fill = row_fill
            notes_cell.border = data_border
            notes_cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Search Queries (column 6) - wrapped text, metadata font
            queries_cell = ws.cell(
                row=row_idx, column=6, value=result_data["search_queries"]
            )
            queries_cell.font = metadata_font
            queries_cell.fill = row_fill
            queries_cell.border = data_border
            queries_cell.alignment = Alignment(wrap_text=True, vertical="top")

            # URL (column 7) - hyperlink in teal (for reference only)
            url_cell = ws.cell(row=row_idx, column=7, value=result_data["url"])
            if result_data["url"]:
                url_cell.hyperlink = result_data["url"]
                url_cell.font = link_font
            else:
                url_cell.font = data_font
            url_cell.fill = row_fill
            url_cell.border = data_border
            url_cell.alignment = Alignment(vertical="top")

            # Search Date (column 8) - centred, date format, metadata-only
            date_cell = ws.cell(row=row_idx, column=8, value=result_data["search_date"])
            date_cell.font = metadata_font
            date_cell.fill = row_fill
            date_cell.border = data_border
            date_cell.alignment = Alignment(horizontal="center", vertical="top")
            if result_data["search_date"]:
                date_cell.number_format = "YYYY-MM-DD"

        # Add data validation dropdowns
        self._add_review_validation(ws, len(results_data))

        # Freeze header row
        ws.freeze_panes = "A2"

        # Add auto-filter to header row
        if results_data:
            ws.auto_filter.ref = f"A1:H{len(results_data) + 1}"

        # Set print settings for Review Results
        self._set_print_settings(ws, orientation="landscape")

        return ws

    def _create_prisma_statistics_sheet(self, wb: Workbook, session_id: str):
        """Create PRISMA Statistics sheet with Agent Grey styling and auto-calculating formulas."""
        ws = wb.create_sheet("PRISMA Statistics")

        # Agent Grey title styling
        title_cell = ws.cell(row=1, column=1, value="PRISMA 2020 Flow Statistics")
        title_cell.font = Font(
            name="Calibri",
            size=16,
            bold=True,
            color=self.COLOUR_DEEP_NAVY,
        )
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:C1")
        ws.row_dimensions[1].height = 30

        # Get session statistics for static values
        results_provider = get_results_provider()
        total_results = results_provider.get_results_count(UUID(session_id))

        # PRISMA 2020 compliant metrics with formulas
        # Column C is Review Decision (3rd column)
        prisma_metrics = [
            ("Records identified", total_results),  # Static value
            ("Duplicate records removed", 0),  # Static placeholder
            (
                "Records screened",
                "=COUNTA('Review Results'!A:A)-1",
            ),  # All results minus header
            (
                "Records excluded",
                "=COUNTIF('Review Results'!C:C,\"Exclude\")",
            ),  # Column C
            (
                "Records included",
                "=COUNTIF('Review Results'!C:C,\"Include\")",
            ),  # Column C
            (
                "Records pending review",
                "=COUNTIF('Review Results'!C:C,\"Maybe\")+COUNTIF('Review Results'!C:C,\"\")",  # Column C
            ),
        ]

        # Table header
        header_row = 3
        for col, header_text in enumerate(["Metric", "Value", "Type"], 1):
            cell = ws.cell(row=header_row, column=col, value=header_text)
            cell.font = Font(
                name="Calibri", size=12, bold=True, color=self.COLOUR_WHITE
            )
            cell.fill = PatternFill(
                start_color=self.COLOUR_DEEP_NAVY,
                end_color=self.COLOUR_DEEP_NAVY,
                fill_type="solid",
            )
            cell.border = self._create_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add metrics with alternating row colours
        off_white_fill = PatternFill(
            start_color=self.COLOUR_OFF_WHITE,
            end_color=self.COLOUR_OFF_WHITE,
            fill_type="solid",
        )
        white_fill = PatternFill(
            start_color=self.COLOUR_WHITE,
            end_color=self.COLOUR_WHITE,
            fill_type="solid",
        )

        for i, (metric, formula_or_value) in enumerate(prisma_metrics, start=4):
            row_fill = off_white_fill if i % 2 == 0 else white_fill

            # Metric name
            metric_cell = ws.cell(row=i, column=1, value=metric)
            metric_cell.font = Font(name="Calibri", size=11, bold=True)
            metric_cell.fill = row_fill
            metric_cell.border = self._create_border()
            metric_cell.alignment = Alignment(horizontal="left", vertical="center")

            # Value or formula
            value_cell = ws.cell(row=i, column=2, value=formula_or_value)
            is_formula = isinstance(
                formula_or_value, str
            ) and formula_or_value.startswith("=")
            if is_formula:
                value_cell.font = Font(
                    name="Calibri", size=11, italic=True, color=self.COLOUR_STATUS_BLUE
                )
            else:
                value_cell.font = Font(name="Calibri", size=11, bold=True)
            value_cell.fill = row_fill
            value_cell.border = self._create_border()
            value_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Type indicator
            type_cell = ws.cell(
                row=i, column=3, value="Formula" if is_formula else "Static"
            )
            type_cell.font = Font(
                name="Calibri", size=10, italic=True, color=self.COLOUR_COOL_GREY_DARK
            )
            type_cell.fill = row_fill
            type_cell.border = self._create_border()
            type_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add explanatory notes with Agent Grey styling
        notes_start_row = len(prisma_metrics) + 6
        notes_header = ws.cell(row=notes_start_row, column=1, value="Usage Notes")
        notes_header.font = Font(
            name="Calibri", size=12, bold=True, color=self.COLOUR_DEEP_NAVY
        )

        notes = [
            "• Formulas automatically update when review decisions change in the Review Results sheet",
            "• 'Records identified' shows the total number of results from search execution",
            "• 'Duplicate records removed' indicates results filtered during processing",
            "• All metrics follow PRISMA 2020 guidelines for systematic review reporting",
        ]

        for i, note in enumerate(notes, start=1):
            note_cell = ws.cell(row=notes_start_row + i, column=1, value=note)
            note_cell.font = Font(
                name="Calibri", size=10, color=self.COLOUR_COOL_GREY_DARK
            )
            note_cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(f"A{notes_start_row + i}:C{notes_start_row + i}")

        # Set column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15

        # Set print settings for PRISMA Statistics
        self._set_print_settings(ws, orientation="portrait")

        return ws

    def _get_results_data(self, session_id: str) -> List[dict]:
        """Get optimized results data with related review decisions."""
        results_provider = get_results_provider()
        results = results_provider.get_results_for_session(UUID(session_id))

        # Get review decisions in bulk to avoid N+1 queries
        decisions = SimpleReviewDecision.objects.filter(
            result__session_id=session_id
        ).select_related("result")

        decision_map = {str(d.result_id): d for d in decisions}

        # Build results data
        results_data = []
        for result in results:  # type: ignore[reportGeneralTypeIssues]
            decision = decision_map.get(str(result.id))

            # Get search queries for this result
            search_queries = self._get_search_queries_for_result(result)

            results_data.append(
                {
                    "title": result.title or "No Title Available",
                    "url": result.url or "",
                    "snippet": result.snippet or "",
                    "search_queries": (
                        ", ".join(search_queries) if search_queries else ""
                    ),
                    "search_date": (
                        result.processed_at.date() if result.processed_at else ""
                    ),
                    "decision": self._format_decision(
                        decision.decision if decision else ""
                    ),
                    "exclusion_reason": (
                        decision.get_exclusion_reason_display()
                        if decision and decision.exclusion_reason
                        else ""
                    ),
                    "notes": decision.notes if decision else "",
                }
            )

        return results_data

    def _get_search_queries_for_result(self, result) -> List[str]:
        """Get all search queries that found this result."""
        queries = []

        # Try to get from raw result -> execution -> search query
        raw_result = getattr(result, "raw_result", None)
        if raw_result:
            execution = getattr(raw_result, "execution", None)
            if execution:
                search_query = getattr(execution, "query", None)
                if search_query:
                    formatted = getattr(search_query, "formatted_query", None)
                    if formatted:
                        queries.append(formatted)
                    elif search_query.query_text:
                        queries.append(search_query.query_text)

        return queries

    def _format_decision(self, decision: str) -> str:
        """Format decision for display in Excel."""
        if not decision:
            return ""
        # Capitalize first letter to match dropdown values
        return decision.capitalize()

    def _add_review_validation(self, ws, num_results: int):
        """Add data validation dropdowns for review decisions and exclusion reasons."""
        if num_results == 0:
            return

        # Review Decision dropdown (column C - 3rd column)
        decision_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(self.REVIEW_DECISIONS)}"',
            allow_blank=True,
        )
        decision_dv.prompt = "Select review decision"
        decision_dv.promptTitle = "Review Decision"
        decision_dv.error = "Please select from the dropdown list"
        decision_dv.errorTitle = "Invalid Selection"

        ws.add_data_validation(decision_dv)
        decision_dv.add(f"C2:C{num_results + 1}")  # Column C (3rd column)

        # Exclusion Reason dropdown (column D - 4th column)
        exclusion_dv = DataValidation(
            type="list",
            formula1=f'"{",".join(self.EXCLUSION_REASONS)}"',
            allow_blank=True,
        )
        exclusion_dv.prompt = "Select reason if excluding result"
        exclusion_dv.promptTitle = "Exclusion Reason"
        exclusion_dv.error = "Please select from the dropdown list"
        exclusion_dv.errorTitle = "Invalid Selection"

        ws.add_data_validation(exclusion_dv)
        exclusion_dv.add(f"D2:D{num_results + 1}")  # Column D (4th column)

    def _create_border(self, style="thin") -> Border:
        """Create a border style using Agent Grey colours."""
        side = Side(border_style=style, color=self.COLOUR_COOL_GREY_LIGHT)
        return Border(left=side, right=side, top=side, bottom=side)

    def _get_decision_fill(self, decision: str, default_fill) -> PatternFill:
        """Get status colour fill based on review decision."""
        if decision == "Include":
            return PatternFill(
                start_color=self.COLOUR_STATUS_GREEN,
                end_color=self.COLOUR_STATUS_GREEN,
                fill_type="solid",
            )
        elif decision == "Exclude":
            return PatternFill(
                start_color=self.COLOUR_ERROR_RED,
                end_color=self.COLOUR_ERROR_RED,
                fill_type="solid",
            )
        elif decision == "Maybe":
            return PatternFill(
                start_color=self.COLOUR_AMBER_GOLD,
                end_color=self.COLOUR_AMBER_GOLD,
                fill_type="solid",
            )
        else:
            # No decision yet - use default alternating row colour
            return default_fill

    def _set_print_settings(self, ws, orientation="landscape"):
        """Configure print settings for professional output."""
        # Page setup
        ws.page_setup.orientation = (
            ws.ORIENTATION_LANDSCAPE
            if orientation == "landscape"
            else ws.ORIENTATION_PORTRAIT
        )
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # Allow multiple pages vertically

        # Margins (in inches)
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3

        # Print options
        ws.print_options.horizontalCentered = True
        ws.print_options.gridLines = False

    def _create_instructions_sheet(self, wb: Workbook, data: dict):
        """Create an Instructions sheet with usage guide."""
        ws = wb.active  # First sheet created
        assert ws is not None
        ws.title = "Instructions"

        # Title
        title_cell = ws.cell(
            row=1, column=1, value="Agent Grey - Offline Backup Instructions"
        )
        title_cell.font = Font(
            name="Calibri", size=18, bold=True, color=self.COLOUR_DEEP_NAVY
        )
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:D1")
        ws.row_dimensions[1].height = 35

        # Export metadata
        session_data = data.get("session_data", {})
        metadata_start = 3
        metadata_items = [
            ("Session Name:", session_data.get("name", "N/A")),
            ("Export Date:", data.get("export_date", "N/A")),
            ("Total Results:", data.get("total_results", "N/A")),
        ]

        for i, (label, value) in enumerate(metadata_items):
            label_cell = ws.cell(row=metadata_start + i, column=1, value=label)
            label_cell.font = Font(
                name="Calibri", size=11, bold=True, color=self.COLOUR_DEEP_NAVY
            )

            value_cell = ws.cell(row=metadata_start + i, column=2, value=value)
            value_cell.font = Font(name="Calibri", size=11)

        # Instructions sections
        instructions_start = metadata_start + len(metadata_items) + 2

        sections = [
            {
                "title": "📋 Overview",
                "content": [
                    "This Excel workbook contains your systematic literature review data for offline analysis.",
                    "All sheets are interconnected - changes in the Review Results sheet "
                    "automatically update PRISMA Statistics.",
                ],
            },
            {
                "title": "📊 Sheet Descriptions",
                "content": [
                    "• Review Results: Main data table with all search results and review decisions",
                    "• PRISMA Statistics: Auto-calculating metrics following PRISMA 2020 guidelines",
                    "• Instructions: This sheet with usage guidance",
                ],
            },
            {
                "title": "✏️ How to Review Results",
                "content": [
                    "1. Navigate to the 'Review Results' sheet",
                    "2. For each result, select a decision from the 'Review Decision' dropdown:",
                    "   • Include: Result meets inclusion criteria",
                    "   • Exclude: Result should be excluded (select reason in next column)",
                    "   • Maybe: Requires further consideration",
                    "3. If excluding, select an 'Exclusion Reason' from the dropdown",
                    "4. Add any relevant notes in the 'Notes' column",
                    "5. Save the file regularly to preserve your work",
                ],
            },
            {
                "title": "🎨 Colour Legend",
                "content": [
                    f"• Green cells: Include decisions (status colour: #{self.COLOUR_STATUS_GREEN})",
                    f"• Amber cells: Maybe decisions (caution colour: #{self.COLOUR_AMBER_GOLD})",
                    f"• Red cells: Exclude decisions (error colour: #{self.COLOUR_ERROR_RED})",
                    "• Alternating row colours improve readability",
                ],
            },
            {
                "title": "💡 Tips",
                "content": [
                    "• Use the auto-filter (arrow icons in header row) to sort and filter results",
                    "• Frozen header row remains visible when scrolling",
                    "• Click on any title or URL to open the source directly in your browser",
                    "• PRISMA Statistics update automatically - no manual calculation needed",
                    "• Print settings are pre-configured for A4 paper",
                ],
            },
        ]

        current_row = instructions_start
        for section in sections:
            # Section title
            title_cell = ws.cell(row=current_row, column=1, value=section["title"])
            title_cell.font = Font(
                name="Calibri", size=13, bold=True, color=self.COLOUR_DEEP_NAVY
            )
            ws.merge_cells(f"A{current_row}:D{current_row}")
            current_row += 1

            # Section content
            for content_line in section["content"]:
                content_cell = ws.cell(row=current_row, column=1, value=content_line)
                content_cell.font = Font(
                    name="Calibri", size=10, color=self.COLOUR_COOL_GREY_DARK
                )
                content_cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.merge_cells(f"A{current_row}:D{current_row}")
                ws.row_dimensions[current_row].height = (
                    30 if len(content_line) > 80 else 20
                )
                current_row += 1

            current_row += 1  # Space between sections

        # Footer
        footer_row = current_row + 2
        footer_cell = ws.cell(
            row=footer_row,
            column=1,
            value="Agent Grey - Precision in Grey Literature Review",
        )
        footer_cell.font = Font(
            name="Calibri", size=10, italic=True, color=self.COLOUR_COOL_GREY_DARK
        )
        footer_cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(f"A{footer_row}:D{footer_row}")

        # Set column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 20

        return ws

    def _save_workbook_to_bytes(self, wb: Workbook) -> bytes:
        """Save workbook to bytes for return."""
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


class ExcelGenerationError(Exception):
    """Custom exception for Excel generation failures."""

    pass
