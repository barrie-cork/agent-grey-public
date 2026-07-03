"""
Demo report generation service for reporting app.

This service handles generation of demo reports when Celery is not available.
"""

import json
from io import BytesIO
from typing import Any

import openpyxl  # type: ignore[reportMissingModuleSource]
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage

from apps.review_results.models import SimpleReviewDecision

from ..models import ExportReport


class DemoReportService:
    """Service for generating demo reports without Celery."""

    def generate(self, report: ExportReport, session: Any) -> None:
        """
        Generate demo report content and save to storage.

        Args:
            report: ExportReport instance to update
            session: SearchSession instance
        """
        format_type = report.export_format

        # Generate content based on format
        if format_type == "csv":
            file_content = self._generate_csv_content(session)
        elif format_type == "json":
            file_content = self._generate_json_content(session)
        elif format_type == "xlsx":
            file_content = self._generate_xlsx_content(session)
        else:
            # Default content for PDF/DOCX
            file_content = self._generate_default_content(session)

        # Save file
        file_name = f"{report.report_type}_{report.id}.{format_type}"
        file_path = f"reports/{session.id}/{file_name}"
        saved_path = default_storage.save(file_path, file_content)

        # Update report metadata
        report.file_path = saved_path
        report.file_name = file_name
        report.file_size_bytes = file_content.size
        report.save(update_fields=["file_path", "file_name", "file_size_bytes"])

        # Mark as completed using status manager
        from .status_manager import ReportStatusManager

        ReportStatusManager.mark_as_completed(report)

    def _generate_csv_content(self, session: Any) -> ContentFile:
        """Generate CSV content with actual session data."""
        content = "Title,URL,Decision,Publication Date,Document Type\n"

        # Get all reviewed results
        decisions = SimpleReviewDecision.objects.filter(session=session).select_related(
            "result"
        )

        for decision in decisions[:100]:  # Limit to first 100 for demo
            result = decision.result
            content += (
                f'"{result.title}","{result.url}","{decision.get_decision_display()}",'
                f'"{result.publication_date or ""}","{result.document_type}"\n'
            )

        return ContentFile(content.encode("utf-8"))

    def _generate_json_content(self, session: Any) -> ContentFile:
        """Generate JSON content with session statistics."""
        from django.utils import timezone

        decisions = SimpleReviewDecision.objects.filter(session=session)
        included = decisions.filter(decision="include").count()
        excluded = decisions.filter(decision="exclude").count()
        maybe = decisions.filter(decision="maybe").count()
        total = decisions.count()

        content = json.dumps(
            {
                "session": {
                    "id": str(session.id),
                    "title": session.title,
                    "description": session.description,
                },
                "results": {
                    "total_reviewed": total,
                    "included": included,
                    "excluded": excluded,
                    "maybe": maybe,
                },
                "generated_at": timezone.now().isoformat(),
            },
            indent=2,
        )

        return ContentFile(content.encode("utf-8"))

    def _generate_xlsx_content(self, session: Any) -> ContentFile:
        """Generate minimal Excel file."""
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None  # Workbook() always creates an active sheet
        ws["A1"] = "Title"
        ws["B1"] = "Status"
        ws["A2"] = session.title
        ws["B2"] = "Completed"

        buffer = BytesIO()
        wb.save(buffer)
        return ContentFile(buffer.getvalue())

    def _generate_default_content(self, session: Any) -> ContentFile:
        """Generate default content for PDF/DOCX formats."""
        content = (
            f"Report for {session.title}\n"
            f"Total Results: 314\n"
            f"Included: 45\n"
            f"Excluded: 269"
        )
        return ContentFile(content.encode("utf-8"))
