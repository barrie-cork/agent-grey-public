"""
Integration tests for Excel import views.

Tests cover:
- Import form display
- File upload and validation
- Session status validation
- Ownership validation
- AJAX request handling
"""

import io
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import Client, TestCase
from django.urls import reverse
from openpyxl import Workbook

from apps.results_manager.models import ProcessedResult
from apps.review_manager.models import SearchSession
from apps.core.tests.utils import create_test_user

User = get_user_model()


class ImportOfflineBackupViewTestCase(TestCase):
    """Test suite for ImportOfflineBackupView."""

    def setUp(self):
        """Set up test fixtures."""
        self.client = Client()

        # Create test user
        self.user = create_test_user()

        # Create test session
        self.session = SearchSession.objects.create(
            owner=self.user, title="Test Session", status="ready_for_review"
        )

        # Create test result
        self.result = ProcessedResult.objects.create(
            session=self.session,
            title="Test Result",
            url="https://example.com/test",
            snippet="Test snippet",
        )

        self.url = reverse("reporting:import_backup", args=[self.session.id])

    def _create_valid_excel_file(self):
        """Create a valid Excel file for testing."""
        from django.core.files.uploadedfile import SimpleUploadedFile

        wb = Workbook()
        ws = wb.create_sheet("Review Results")

        # Add headers
        headers = [
            "Title",
            "Snippet",
            "Search Date",
            "Review Decision",
            "Exclusion Reason",
            "Notes",
            "Search Query(ies)",
            "URL",
        ]
        ws.append(headers)

        # Add data row matching our test result
        ws.append(
            [
                "Test Result",
                "Test snippet",
                "2025-10-17",
                "Include",
                "",
                "Test note",
                "test query",
                "https://example.com/test",
            ]
        )

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Return as SimpleUploadedFile with proper name
        return SimpleUploadedFile(
            "test_backup.xlsx",
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_get_import_form_unauthenticated(self):
        """Test GET redirects to login for unauthenticated users."""
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 302)
        self.assertIn("/login/", response.url)

    def test_get_import_form_authenticated(self):
        """Test GET displays import form for authenticated users."""
        self.client.login(username=self.user.username, password="testpass123")

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "reporting/import_backup.html")
        self.assertIn("form", response.context)
        self.assertIn("session", response.context)

    def test_get_import_form_wrong_user(self):
        """Test GET returns 404 for session not owned by user."""
        # Create another user
        other_user = create_test_user(username_prefix="otheruser")

        self.client.login(username=other_user.username, password="testpass123")

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 404)

    def test_get_import_form_wrong_session_status(self):
        """Test GET redirects for session not in review phase."""
        # Change session status
        self.session.status = "draft"
        self.session.save()

        self.client.login(username=self.user.username, password="testpass123")

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 302)
        self.assertRedirects(
            response, reverse("review_results:overview", args=[self.session.id])
        )

    @patch("apps.reporting.tasks.import_excel_backup_task")
    def test_post_import_valid_file(self, mock_task):
        """Test POST with valid Excel file queues import task."""
        self.client.login(username=self.user.username, password="testpass123")

        # Mock Celery task
        mock_task.delay.return_value.id = "test-task-id"

        excel_file = self._create_valid_excel_file()

        response = self.client.post(
            self.url, {"excel_file": excel_file, "session": str(self.session.id)}
        )

        # Should redirect to review results
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(
            response, reverse("review_results:overview", args=[self.session.id])
        )

        # Verify task was queued
        self.assertTrue(mock_task.delay.called)

    @patch("apps.reporting.tasks.import_excel_backup_task")
    def test_post_import_ajax_request(self, mock_task):
        """Test POST with AJAX returns JSON response."""
        self.client.login(username=self.user.username, password="testpass123")

        # Mock Celery task
        mock_task.delay.return_value.id = "test-task-id"

        excel_file = self._create_valid_excel_file()

        response = self.client.post(
            self.url,
            {"excel_file": excel_file, "session": str(self.session.id)},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        # Should return JSON
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/json")

        data = response.json()
        self.assertEqual(data["status"], "processing")
        self.assertEqual(data["task_id"], "test-task-id")

    def test_post_import_invalid_file_extension(self):
        """Test POST with wrong file extension fails validation."""
        self.client.login(username=self.user.username, password="testpass123")

        # Create text file instead of Excel
        text_file = io.BytesIO(b"This is not an Excel file")
        text_file.name = "test.txt"

        response = self.client.post(
            self.url, {"excel_file": text_file, "session": str(self.session.id)}
        )

        # Should show form with errors
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "reporting/import_backup.html")
        # Check that form has errors
        self.assertTrue(response.context["form"].errors)
        self.assertIn("excel_file", response.context["form"].errors)
        # Verify error message content
        error_message = str(response.context["form"].errors["excel_file"])
        self.assertIn("not allowed", error_message.lower())

    def test_post_import_file_too_large(self):
        """Test POST with oversized file fails validation."""
        self.client.login(username=self.user.username, password="testpass123")

        # Create file larger than 10MB
        large_file = io.BytesIO(b"x" * (11 * 1024 * 1024))
        large_file.name = "large.xlsx"

        response = self.client.post(
            self.url, {"excel_file": large_file, "session": str(self.session.id)}
        )

        # Should show form with errors
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "reporting/import_backup.html")
        # Form should have error about file size

    def test_post_import_wrong_session_status(self):
        """Test POST fails for session not in review phase."""
        # Change session status
        self.session.status = "draft"
        self.session.save()

        self.client.login(username=self.user.username, password="testpass123")

        excel_file = self._create_valid_excel_file()

        response = self.client.post(
            self.url, {"excel_file": excel_file, "session": str(self.session.id)}
        )

        # Should redirect with error message
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(
            response, reverse("review_results:overview", args=[self.session.id])
        )

    def test_post_import_missing_file(self):
        """Test POST without file fails validation."""
        self.client.login(username=self.user.username, password="testpass123")

        response = self.client.post(self.url, {"session": str(self.session.id)})

        # Should show form with errors
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "reporting/import_backup.html")
        self.assertFormError(
            response.context["form"], "excel_file", "This field is required."
        )


class ImportButtonDisplayTestCase(TestCase):
    """Test suite for import button display on review results page."""

    def setUp(self):
        """Set up test fixtures."""
        self.client = Client()

        # Create test user
        self.user = create_test_user()

        # Create test session
        self.session = SearchSession.objects.create(
            owner=self.user, title="Test Session", status="ready_for_review"
        )

        # Create test result
        self.result = ProcessedResult.objects.create(
            session=self.session,
            title="Test Result",
            url="https://example.com/test",
            snippet="Test snippet",
        )

        self.url = reverse("review_results:overview", args=[self.session.id])

    def test_import_button_visible_ready_for_review(self):
        """Test import button is visible when session is ready for review."""
        self.session.status = "ready_for_review"
        self.session.save()

        self.client.login(username=self.user.username, password="testpass123")

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Import Changes")
        self.assertContains(
            response, reverse("reporting:import_backup", args=[self.session.id])
        )

    def test_import_button_visible_under_review(self):
        """Test import button is visible when session is under review."""
        self.session.status = "under_review"
        self.session.save()

        self.client.login(username=self.user.username, password="testpass123")

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Import Changes")

    def test_import_button_not_visible_wrong_status(self):
        """Test import button is not visible for other session statuses."""
        self.session.status = "draft"
        self.session.save()

        self.client.login(username=self.user.username, password="testpass123")

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Import Changes")
