"""
Unit tests for Excel import service.

Tests cover:
- Excel file parsing and validation
- Schema validation
- Change detection
- Database synchronisation
- Error handling
"""

import io
from unittest.mock import patch

from django.core.exceptions import ValidationError
from django.test import TestCase
from openpyxl import Workbook

from apps.reporting.services.excel_import_service import (
    ExcelImportError,
    ExcelImportService,
)
from apps.core.tests.utils import create_test_user
from apps.results_manager.models import ProcessedResult
from apps.review_manager.models import SearchSession
from apps.review_results.models import SimpleReviewDecision


class ExcelImportServiceTestCase(TestCase):
    """Test suite for ExcelImportService."""

    def setUp(self):
        """Set up test fixtures."""
        self.service = ExcelImportService()

        # Create test user
        self.user = create_test_user()

        # Create test session
        self.session = SearchSession.objects.create(
            owner=self.user, title="Test Session", status="ready_for_review"
        )

    def _create_valid_workbook(self):
        """Create a valid Excel workbook for testing."""
        wb = Workbook()
        ws = wb.create_sheet("Review Results")

        # Add headers
        headers = [
            "Title",
            "Snippet",
            "Search Date",
            "Review Decision",
            "Exclusion Reason",
            "Notes",
            "Search Query(ies)",
            "URL",
        ]
        ws.append(headers)

        # Add sample data row
        ws.append(
            [
                "Test Result",
                "Test snippet",
                "2025-10-17",
                "Include",
                "",
                "Test note",
                "test query",
                "https://example.com/test",
            ]
        )

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        return wb

    def _workbook_to_bytes(self, wb):
        """Convert workbook to bytes for file-like object."""
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def test_validate_workbook_structure_valid(self):
        """Test validation of valid workbook structure."""
        wb = self._create_valid_workbook()

        # Should not raise exception
        self.service._validate_workbook_structure(wb)

    def test_validate_workbook_structure_missing_sheet(self):
        """Test validation fails for missing Review Results sheet."""
        wb = Workbook()

        with self.assertRaises(ValidationError) as cm:
            self.service._validate_workbook_structure(wb)

        self.assertIn("Missing 'Review Results' sheet", str(cm.exception))

    def test_validate_workbook_structure_wrong_headers(self):
        """Test validation fails for incorrect headers."""
        wb = Workbook()
        ws = wb.create_sheet("Review Results")

        # Add wrong headers
        ws.append(["Wrong", "Headers", "Here"])

        with self.assertRaises(ValidationError) as cm:
            self.service._validate_workbook_structure(wb)

        self.assertIn("Column headers don't match", str(cm.exception))

    def test_extract_review_data_valid(self):
        """Test extracting review data from valid workbook."""
        wb = self._create_valid_workbook()

        review_data = self.service._extract_review_data(wb)

        self.assertEqual(len(review_data), 1)
        self.assertEqual(review_data[0]["title"], "Test Result")
        self.assertEqual(review_data[0]["decision"], "include")  # lowercase
        self.assertEqual(review_data[0]["notes"], "Test note")

    def test_extract_review_data_skips_empty_rows(self):
        """Test that empty rows are skipped during extraction."""
        wb = self._create_valid_workbook()
        ws = wb["Review Results"]

        # Add empty row
        ws.append(["", "", "", "", "", "", "", ""])

        review_data = self.service._extract_review_data(wb)

        # Should only have 1 row (empty row skipped)
        self.assertEqual(len(review_data), 1)

    def test_extract_review_data_normalizes_decision(self):
        """Test that decision values are normalized to lowercase."""
        wb = self._create_valid_workbook()
        ws = wb["Review Results"]

        # Add row with capitalized decision
        ws.append(
            [
                "Test Result 2",
                "Test snippet",
                "2025-10-17",
                "EXCLUDE",  # All caps
                "Not relevant",
                "",
                "test query",
                "https://example.com/test2",
            ]
        )

        review_data = self.service._extract_review_data(wb)

        self.assertEqual(review_data[1]["decision"], "exclude")  # lowercase

    def test_extract_review_data_handles_invalid_decision(self):
        """Test that invalid decisions are set to empty string."""
        wb = self._create_valid_workbook()
        ws = wb["Review Results"]

        # Add row with invalid decision
        ws.append(
            [
                "Test Result 2",
                "Test snippet",
                "2025-10-17",
                "InvalidDecision",  # Invalid
                "",
                "",
                "test query",
                "https://example.com/test2",
            ]
        )

        review_data = self.service._extract_review_data(wb)

        # Invalid decision should be normalized to empty string
        self.assertEqual(review_data[1]["decision"], "")

    def test_find_matching_row_success(self):
        """Test finding matching row by title and URL."""
        # Create test result
        result = ProcessedResult.objects.create(
            session=self.session,
            title="Test Result",
            url="https://example.com/test",
            snippet="Test snippet",
        )

        review_data = [
            {
                "title": "Test Result",
                "url": "https://example.com/test",
                "decision": "include",
                "exclusion_reason": "",
                "notes": "Test note",
            }
        ]

        match = self.service._find_matching_row(result, review_data)

        self.assertIsNotNone(match)
        self.assertEqual(match["title"], "Test Result")

    def test_find_matching_row_no_match(self):
        """Test no match found when title/URL don't match."""
        # Create test result
        result = ProcessedResult.objects.create(
            session=self.session,
            title="Different Title",
            url="https://example.com/different",
            snippet="Test snippet",
        )

        review_data = [
            {
                "title": "Test Result",
                "url": "https://example.com/test",
                "decision": "include",
                "exclusion_reason": "",
                "notes": "Test note",
            }
        ]

        match = self.service._find_matching_row(result, review_data)

        self.assertIsNone(match)

    @patch("apps.reporting.services.excel_import_service.get_results_provider")
    def test_detect_changes_new_decision(self, mock_provider):
        """Test detecting changes for new review decision."""
        # Create test result
        result = ProcessedResult.objects.create(
            session=self.session,
            title="Test Result",
            url="https://example.com/test",
            snippet="Test snippet",
        )

        # Mock provider
        mock_provider.return_value.get_results_for_session.return_value = [result]

        review_data = [
            {
                "title": "Test Result",
                "url": "https://example.com/test",
                "decision": "include",
                "exclusion_reason": "",
                "notes": "Test note",
            }
        ]

        changes = self.service._detect_changes(str(self.session.id), review_data)

        self.assertEqual(len(changes), 1)
        self.assertEqual(changes[0]["result_id"], result.id)
        self.assertEqual(changes[0]["old_values"]["decision"], "")
        self.assertEqual(changes[0]["new_values"]["decision"], "include")

    @patch("apps.reporting.services.excel_import_service.get_results_provider")
    def test_detect_changes_modified_decision(self, mock_provider):
        """Test detecting changes when decision is modified."""
        # Create test result with existing decision
        result = ProcessedResult.objects.create(
            session=self.session,
            title="Test Result",
            url="https://example.com/test",
            snippet="Test snippet",
        )

        SimpleReviewDecision.objects.create(
            result=result,
            session=self.session,
            reviewer=self.user,
            decision="maybe",
            notes="Old note",
        )

        # Mock provider
        mock_provider.return_value.get_results_for_session.return_value = [result]

        review_data = [
            {
                "title": "Test Result",
                "url": "https://example.com/test",
                "decision": "include",
                "exclusion_reason": "",
                "notes": "New note",
            }
        ]

        changes = self.service._detect_changes(str(self.session.id), review_data)

        self.assertEqual(len(changes), 1)
        self.assertEqual(changes[0]["old_values"]["decision"], "maybe")
        self.assertEqual(changes[0]["new_values"]["decision"], "include")
        self.assertEqual(changes[0]["old_values"]["notes"], "Old note")
        self.assertEqual(changes[0]["new_values"]["notes"], "New note")

    @patch("apps.reporting.services.excel_import_service.get_results_provider")
    def test_detect_changes_no_changes(self, mock_provider):
        """Test detecting changes when values are unchanged."""
        # Create test result with existing decision
        result = ProcessedResult.objects.create(
            session=self.session,
            title="Test Result",
            url="https://example.com/test",
            snippet="Test snippet",
        )

        SimpleReviewDecision.objects.create(
            result=result,
            session=self.session,
            reviewer=self.user,
            decision="include",
            notes="Same note",
        )

        # Mock provider
        mock_provider.return_value.get_results_for_session.return_value = [result]

        review_data = [
            {
                "title": "Test Result",
                "url": "https://example.com/test",
                "decision": "include",
                "exclusion_reason": "",
                "notes": "Same note",
            }
        ]

        changes = self.service._detect_changes(str(self.session.id), review_data)

        # No changes detected
        self.assertEqual(len(changes), 0)

    def test_apply_changes_creates_new_decision(self):
        """Test applying changes creates new review decision."""
        # Create test result
        result = ProcessedResult.objects.create(
            session=self.session,
            title="Test Result",
            url="https://example.com/test",
            snippet="Test snippet",
        )

        changes = [
            {
                "result_id": result.id,
                "old_values": {"decision": "", "exclusion_reason": "", "notes": ""},
                "new_values": {
                    "decision": "include",
                    "exclusion_reason": "",
                    "notes": "Test note",
                },
            }
        ]

        summary = self.service._apply_changes(
            str(self.session.id), changes, reviewer=self.user
        )

        self.assertEqual(summary["total_changes"], 1)
        self.assertEqual(summary["updated"], 1)
        self.assertEqual(summary["errors"], 0)

        # Verify decision was created
        decision = SimpleReviewDecision.objects.get(result=result)
        self.assertEqual(decision.decision, "include")
        self.assertEqual(decision.notes, "Test note")
        self.assertEqual(decision.session, self.session)
        self.assertEqual(decision.reviewer, self.user)

    def test_apply_changes_updates_existing_decision(self):
        """Test applying changes updates existing review decision."""
        # Create test result with existing decision
        result = ProcessedResult.objects.create(
            session=self.session,
            title="Test Result",
            url="https://example.com/test",
            snippet="Test snippet",
        )

        decision = SimpleReviewDecision.objects.create(
            result=result,
            session=self.session,
            reviewer=self.user,
            decision="maybe",
            notes="Old note",
        )

        changes = [
            {
                "result_id": result.id,
                "old_values": {
                    "decision": "maybe",
                    "exclusion_reason": "",
                    "notes": "Old note",
                },
                "new_values": {
                    "decision": "include",
                    "exclusion_reason": "",
                    "notes": "New note",
                },
            }
        ]

        summary = self.service._apply_changes(
            str(self.session.id), changes, reviewer=self.user
        )

        self.assertEqual(summary["total_changes"], 1)
        self.assertEqual(summary["updated"], 1)
        self.assertEqual(summary["errors"], 0)

        # Verify decision was updated
        decision.refresh_from_db()
        self.assertEqual(decision.decision, "include")
        self.assertEqual(decision.notes, "New note")

    @patch("apps.reporting.services.excel_import_service.load_workbook")
    @patch("apps.reporting.services.excel_import_service.get_results_provider")
    def test_import_excel_backup_success(self, mock_provider, mock_load_wb):
        """Test successful Excel import end-to-end."""
        # Create test result
        result = ProcessedResult.objects.create(
            session=self.session,
            title="Test Result",
            url="https://example.com/test",
            snippet="Test snippet",
        )

        # Mock workbook
        wb = self._create_valid_workbook()
        mock_load_wb.return_value = wb

        # Mock provider
        mock_provider.return_value.get_results_for_session.return_value = [result]

        # Create file-like object
        excel_file = io.BytesIO(b"fake excel content")

        summary = self.service.import_excel_backup(
            str(self.session.id), excel_file, reviewer=self.user
        )

        self.assertEqual(summary["total_changes"], 1)
        self.assertEqual(summary["updated"], 1)
        self.assertEqual(summary["errors"], 0)

    @patch("apps.reporting.services.excel_import_service.load_workbook")
    def test_import_excel_backup_validation_error(self, mock_load_wb):
        """Test Excel import with validation error."""
        # Mock workbook with missing sheet
        wb = Workbook()
        mock_load_wb.return_value = wb

        excel_file = io.BytesIO(b"fake excel content")

        with self.assertRaises(ExcelImportError) as cm:
            self.service.import_excel_backup(str(self.session.id), excel_file)

        self.assertIn("validation failed", str(cm.exception))
